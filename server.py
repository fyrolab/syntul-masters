#!/usr/bin/env python3
"""
Сынтул-Мастерс 2026 — сервер регистрации.
Запуск: python3 server.py
Откроется на http://localhost:8080
Админка:  http://localhost:8080/admin.html  (пароль: syntul2026admin)
"""

import http.server
import json
import os
import uuid
from datetime import datetime
from html import escape
from urllib.parse import parse_qs, urlparse
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load .env file if exists
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())

PORT = int(os.environ.get("PORT", "8080"))
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
DATA_FILE = os.path.join(DATA_DIR, "registrations.json")

ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "changeme")

VALID_GENDERS = {"М", "Ж"}
VALID_BOATS = {"К-1", "К-2", "К-4", "С-1", "С-2", "С-4"}
VALID_DISTANCES = {"200 м", "500 м", "1000 м", "5000 м", "Эстафета 4×200"}


def read_registrations():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        content = f.read()
    if not content.strip():
        return []
    data = json.loads(content)
    return data if isinstance(data, list) else []


def save_registrations(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def sanitize(value):
    return escape(str(value).strip())


def make_xlsx(registrations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = [
        "№", "Дата заявки", "Фамилия", "Имя", "Отчество",
        "Дата рождения", "Пол", "Страна", "Город",
        "Звание", "Команда", "Класс лодки", "Телефон", "Дистанции",
    ]

    # Column widths
    col_widths = [5, 14, 18, 14, 16, 14, 7, 14, 16, 12, 20, 13, 20, 30]

    # Styles
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0A2463", end_color="0A2463", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D8DFE9"),
        right=Side(style="thin", color="D8DFE9"),
        top=Side(style="thin", color="D8DFE9"),
        bottom=Side(style="thin", color="D8DFE9"),
    )
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=False)
    center_align = Alignment(horizontal="center", vertical="center")
    even_fill = PatternFill(start_color="E8F1FB", end_color="E8F1FB", fill_type="solid")

    # Write headers
    for col_idx, (title, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Write data
    center_cols = {1, 7}  # №, Пол
    for i, reg in enumerate(registrations, 1):
        row_num = i + 1
        row_data = [
            i,
            reg.get("timestamp", ""),
            reg.get("surname", ""),
            reg.get("firstname", ""),
            reg.get("patronymic", ""),
            reg.get("birthdate", ""),
            reg.get("gender", ""),
            reg.get("country", ""),
            reg.get("city", ""),
            reg.get("rank", ""),
            reg.get("team", ""),
            reg.get("boat_class", ""),
            reg.get("phone", ""),
            ", ".join(reg.get("distances", [])),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in center_cols:
                cell.alignment = center_align
            else:
                cell.alignment = data_align
            # Zebra striping
            if i % 2 == 0:
                cell.fill = even_fill

    # Auto-filter on all columns
    last_col = get_column_letter(len(headers))
    last_row = max(len(registrations) + 1, 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


class Handler(http.server.SimpleHTTPRequestHandler):

    def send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _check_admin_token(self):
        """Check admin token from X-Admin-Token header."""
        token = self.headers.get("X-Admin-Token", "")
        return token == ADMIN_TOKEN

    def do_GET(self):
        parsed = urlparse(self.path)

        # Admin: list registrations (token in header)
        if parsed.path == "/api":
            if not self._check_admin_token():
                self.send_json({"success": False, "error": "Неверный токен"}, 401)
                return

            registrations = read_registrations()
            self.send_json({
                "success": True,
                "count": len(registrations),
                "data": registrations,
            })
            return

        # Serve static files
        super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        content_length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(content_length).decode("utf-8") if content_length else ""

        # ─── Admin: Excel export ───
        if parsed.path == "/api/export":
            if not self._check_admin_token():
                self.send_json({"success": False, "error": "Неверный токен"}, 401)
                return

            registrations = read_registrations()
            xlsx_data = make_xlsx(registrations)
            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header(
                "Content-Disposition",
                'attachment; filename="syntul_masters_2026.xlsx"',
            )
            self.send_header("Content-Length", str(len(xlsx_data)))
            self.end_headers()
            self.wfile.write(xlsx_data)
            return

        # ─── Public: new registration ───
        if parsed.path != "/api":
            self.send_json({"success": False, "error": "Not found"}, 404)
            return

        try:
            body = json.loads(raw)
        except (json.JSONDecodeError, ValueError):
            self.send_json({"success": False, "error": "Некорректные данные"}, 400)
            return

        # Validate required fields
        required = ["surname", "firstname", "birthdate", "gender", "country", "city", "team", "boat_class", "phone"]
        for field in required:
            if not body.get(field, "").strip():
                self.send_json({"success": False, "error": f"Поле '{field}' обязательно"}, 400)
                return

        if body["gender"] not in VALID_GENDERS:
            self.send_json({"success": False, "error": "Некорректный пол"}, 400)
            return

        if body["boat_class"] not in VALID_BOATS:
            self.send_json({"success": False, "error": "Некорректный класс лодки"}, 400)
            return

        distances = []
        for d in body.get("distances", []):
            if d in VALID_DISTANCES:
                distances.append(d)

        entry = {
            "id": f"reg_{uuid.uuid4().hex[:12]}",
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "surname": sanitize(body["surname"]),
            "firstname": sanitize(body["firstname"]),
            "patronymic": sanitize(body.get("patronymic", "")),
            "birthdate": sanitize(body["birthdate"]),
            "gender": body["gender"],
            "country": sanitize(body["country"]),
            "city": sanitize(body["city"]),
            "rank": sanitize(body.get("rank", "")),
            "team": sanitize(body["team"]),
            "boat_class": body["boat_class"],
            "phone": sanitize(body["phone"]),
            "distances": distances,
        }

        registrations = read_registrations()
        registrations.append(entry)
        save_registrations(registrations)

        self.send_json({"success": True, "id": entry["id"]})

    def log_message(self, format, *args):
        msg = format % args
        print(f"  {msg}")


def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    os.makedirs(DATA_DIR, exist_ok=True)

    server = http.server.HTTPServer(("", PORT), Handler)
    print(f"\n  Сынтул-Мастерс 2026 — сервер запущен")
    print(f"  ────────────────────────────────────")
    print(f"  Регистрация:  http://localhost:{PORT}")
    print(f"  Админ-панель: http://localhost:{PORT}/admin.html")
    print(f"  Пароль админа: {'*' * len(ADMIN_TOKEN)}")
    print(f"  ────────────────────────────────────")
    print(f"  Ctrl+C для остановки\n")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Сервер остановлен.")
        server.server_close()


if __name__ == "__main__":
    main()
